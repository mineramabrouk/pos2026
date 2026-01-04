from .models import Product, StockMovement, ExchangeRate, Category, Sale, SaleItem, CashTransaction
from .forms import StockMovementForm, ExchangeRateForm, CategoryForm, CashTransactionForm
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, DetailView
from django.urls import reverse_lazy
from django.http import JsonResponse
from django.db import transaction
from django.contrib.auth.decorators import login_required
import json

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.groups.filter(name='Admin').exists()

class ProductListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    model = Product
    template_name = 'sales/product_list.html'
    context_object_name = 'products'
    ordering = ['name']

class ProductCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = Product
    template_name = 'sales/product_form.html'
    fields = ['name', 'category', 'price', 'cost', 'stock', 'barcode', 'image']
    success_url = reverse_lazy('product_list')

class ProductUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = Product
    template_name = 'sales/product_form.html'
    fields = ['name', 'category', 'price', 'cost', 'stock', 'barcode', 'image']
    success_url = reverse_lazy('product_list')

class ProductDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Product
    template_name = 'sales/product_confirm_delete.html'
    success_url = reverse_lazy('product_list')

class StockMovementCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = StockMovement
    form_class = StockMovementForm
    template_name = 'sales/stock_movement_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.product = get_object_or_404(Product, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.product = self.product
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.product
        return context

    def get_success_url(self):
        return reverse_lazy('product_list')

class ExchangeRateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'sales/exchange_rate.html'
    success_url = reverse_lazy('exchange_rate')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_rate'] = ExchangeRate.objects.order_by('-date_set').first()
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        new_rate = form.instance.rate
        
        # Bulk update product prices
        products_updated = 0
        for product in Product.objects.all():
            if product.price_usd:
                product.price = product.price_usd * new_rate
                product.save()
                products_updated += 1
        
        messages.success(self.request, f'Tipo de cambio actualizado a {new_rate}. Se recalcularon los precios de {products_updated} productos.')
        return response

class CategoryListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    model = Category
    template_name = 'sales/category_list.html'
    context_object_name = 'categories'
    ordering = ['name']

class CategoryCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'sales/category_form.html'
    success_url = reverse_lazy('category_list')

class CategoryUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'sales/category_form.html'
    success_url = reverse_lazy('category_list')

class CategoryDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Category
    template_name = 'sales/category_confirm_delete.html'
    success_url = reverse_lazy('category_list')

class POSView(LoginRequiredMixin, TemplateView):
    template_name = 'sales/pos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Serialize products for JS
        products = Product.objects.filter(stock__gt=0).select_related('category')
        products_data = []
        for p in products:
            products_data.append({
                'id': p.id,
                'name': p.name,
                'price': float(p.price),
                'stock': p.stock,
                'category': p.category.name if p.category else 'Sin Categoría',
                'barcode': p.barcode,
                'image_url': p.image.url if p.image else '',
            })
        context['products_json'] = json.dumps(products_data)
        return context

@login_required
def create_sale(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            
            if not items:
                return JsonResponse({'success': False, 'error': 'El carrito está vacío.'})

            with transaction.atomic():
                # Create Sale
                sale = Sale.objects.create(
                    salesperson=request.user,
                    total_amount=0 # Will calculate
                )
                
                total_amount = 0
                
                for item in items:
                    product_id = item.get('id')
                    quantity = int(item.get('quantity', 0))
                    price = float(item.get('price', 0)) # Editable price
                    
                    if quantity <= 0:
                        continue
                        
                    product = Product.objects.select_for_update().get(pk=product_id)
                    
                    if product.stock < quantity:
                        raise ValueError(f"Stock insuficiente para {product.name}. Disponible: {product.stock}")
                    
                    # Deduct stock
                    product.stock -= quantity
                    product.save()
                    
                    # Create SaleItem
                    SaleItem.objects.create(
                        sale=sale,
                        product=product,
                        quantity=quantity,
                        price=price
                    )
                    
                    total_amount += price * quantity
                
                sale.total_amount = total_amount
                sale.save()
                
            return JsonResponse({'success': True, 'sale_id': sale.receipt_number})
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
            
    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

class ReceiptView(LoginRequiredMixin, DetailView):
    model = Sale
    template_name = 'sales/receipt.html'
    context_object_name = 'sale'
    slug_field = 'receipt_number'
    slug_url_kwarg = 'receipt_number'

class CashTransactionCreateView(LoginRequiredMixin, CreateView):
    model = CashTransaction
    form_class = CashTransactionForm
    template_name = 'sales/cash_transaction.html'
    success_url = reverse_lazy('dashboard') # Redirect to Dashboard after adding

    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(self.request, 'Movimiento de caja registrado correctamente.')
        return super().form_valid(form)

class ImportProductsView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    template_name = 'sales/import_products.html'

    def post(self, request, *args, **kwargs):
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Por favor seleccione un archivo.')
            return redirect('import_products')

        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un Excel (.xlsx).')
            return redirect('import_products')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Map headers to expected fields (case insensitive)
            header_map = {}
            for i, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'nombre' in header_lower: header_map['name'] = i
                    elif 'categor' in header_lower: header_map['category'] = i
                    elif 'precio' in header_lower: header_map['price'] = i
                    elif 'costo' in header_lower: header_map['cost'] = i
                    elif 'stock' in header_lower: header_map['stock'] = i
                    elif 'código' in header_lower or 'codigo' in header_lower or 'barcode' in header_lower: header_map['barcode'] = i

            # Validate required fields
            required_fields = ['name', 'price', 'cost']
            missing_fields = [field for field in required_fields if field not in header_map]
            
            if missing_fields:
                messages.error(request, f'Faltan columnas obligatorias: {", ".join(missing_fields)}')
                return redirect('import_products')

            products_created = 0
            products_updated = 0
            
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[header_map['name']]: # Skip empty rows
                        continue

                    # Extract values
                    name = row[header_map['name']]
                    price = row[header_map['price']]
                    cost = row[header_map['cost']]
                    
                    category_name = row[header_map['category']] if 'category' in header_map and row[header_map['category']] else None
                    stock = row[header_map['stock']] if 'stock' in header_map and row[header_map['stock']] is not None else 0
                    barcode = str(row[header_map['barcode']]) if 'barcode' in header_map and row[header_map['barcode']] else None

                    # Get or create category
                    category = None
                    if category_name:
                        category, _ = Category.objects.get_or_create(name=category_name)

                    # Check if product exists by barcode (if provided) or name
                    product = None
                    if barcode:
                        product = Product.objects.filter(barcode=barcode).first()
                    
                    if not product:
                        product = Product.objects.filter(name=name).first()

                    if product:
                        # Update existing
                        product.name = name
                        product.category = category
                        product.price = price
                        product.cost = cost
                        product.stock = stock # Optional: maybe add to stock instead of replace? For now replace.
                        if barcode: product.barcode = barcode
                        product.save()
                        products_updated += 1
                    else:
                        # Create new
                        Product.objects.create(
                            name=name,
                            category=category,
                            price=price,
                            cost=cost,
                            stock=stock,
                            barcode=barcode
                        )
                        products_created += 1

            messages.success(request, f'Importación completada: {products_created} creados, {products_updated} actualizados.')
            return redirect('product_list')

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('import_products')

